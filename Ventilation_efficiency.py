# -*- coding: utf-8 -*-
"""
Created on Wed Jul 22 11:19:49 2020
trying to automte everything if problem go to version 14
@author: Devineni
"""

import pandas as pd
import numpy as np
import statistics
from statistics import mean
import time
import datetime as dt
import matplotlib.pyplot as plt
import operator # for plotting

from openpyxl import load_workbook

# import mysql.connector
import os
import pymysql
from sqlalchemy import create_engine

from easygui import *
import sys

#from recalibration import clean_sql_reg   

def prRed(skk): print("\033[31;1;m {}\033[00m" .format(skk)) 
def prYellow(skk): print("\033[33;1;m {}\033[00m" .format(skk)) 



import warnings
warnings.filterwarnings('ignore')


from uncertainties import ufloat
from uncertainties import unumpy
#%% control plot properties
import datetime
import matplotlib.dates as mdates

import matplotlib.units as munits
from pylab import rcParams
rcParams['figure.figsize'] = 7,4.5
plt.rcParams["font.family"] = "calibri"
plt.rcParams["font.weight"] = "normal"
plt.rcParams["font.size"] = 10

#%% Selection Message box
'''
This section deals with taking input selection of the experiment
easygui module was used to create the dialogue boxes for easy input
this is just a more visual way for experiment selection
'''

msg ="Please select a Location/Season you like to analyze"
title = "Season selection"
choices = ["ESHL_summer", "ESHL_winter", "CBo_summer", "CBo_winter"]
database = choicebox(msg, title, choices)


times = pd.read_excel('C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/excel_files/Times_thesis.xlsx', sheet_name= database)


for experiment in times['short name']:
    z = int(times[times['short name'] == experiment].index.values)
    Vdot_sheets = {"ESHL_summer":"ESHL_Vdot", "ESHL_winter":"ESHL_Vdot", "CBo_summer":"CBo_Vdot", "CBo_winter":"CBo_Vdot"}
#%%
    '''
    This code extracts the correct value of Vdot volume in m3 hr-1 required for 
    the experiment based on svens paper reference to be updated
    for now this is only for intensive ventilation. This is to be further automated
    for entire selection
    '''
    if Vdot_sheets[database] == "ESHL_Vdot":

        Vdot = pd.read_excel("C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/excel_files/Vdot_results.xlsx", sheet_name = Vdot_sheets[database])
        
        Kü_20_sup = ufloat(Vdot.at[0,'Vdot_sup'], Vdot.at[0,'Vdot_sup_uncertainity'])
        
        d = {}
        
        for i in range(len(Vdot)):
            d[Vdot.loc[i,"Level"] + "_sup"] = ufloat(Vdot.at[i,'Vdot_sup'], Vdot.at[i,'Vdot_sup_uncertainity'])
            d[Vdot.loc[i,"Level"] + "_exh"] = ufloat(-Vdot.at[i,'Vdot_exh'], Vdot.at[i,'Vdot_exh_uncertainity'])
        
        x = pd.DataFrame(d, index=["Vdot"])
        x = x.transpose()
        
        '''
        Form the results excel sheet, the Vdot required for out experiment is calculated below
        '''
        
        
        i = z
        
        wz = int(times.at[i,'Volume flow (SZ, WZ)'])
        bd = int(times.at[i,'Volume flow (BD)'])
        ku = int(times.at[i,'Volume flow (Kü)'])
        ku_ex = int(times.at[i,'Volume flow (Kü_exhaust)'])
        print(times.at[i,'Name'])
        
        
        a = max(x.at['SZ02_'+ str(wz) +'_sup','Vdot'] + x.at['WZ_'+ str(wz) +'_sup','Vdot'] + x.at['BD_'+ str(bd) +'_sup','Vdot'], x.at['SZ01_'+ str(wz) +'_exh','Vdot'] + x.at['Kü_'+ str(ku) +'_exh','Vdot'] + x.at['BD_'+ str(bd) +'_exh','Vdot'] + x.at['Kü_Ex_'+ str(ku_ex) +'_exh','Vdot']) #
        
        b = max(x.at['SZ02_'+ str(wz) +'_exh','Vdot'] + x.at['WZ_'+ str(wz) +'_exh','Vdot'] + x.at['BD_'+ str(bd) +'_exh','Vdot'] + x.at['Kü_Ex_'+ str(ku_ex) +'_exh','Vdot'], x.at['SZ01_'+ str(wz) +'_sup','Vdot'] + x.at['Kü_'+ str(ku) +'_sup','Vdot'] + x.at['BD_'+ str(bd) +'_sup','Vdot'])
        
        Vdot_imported = (a+b)/2
        
        print(f"Vdot = {Vdot_imported}\n")
        level = "wz={wz}_ku={ku}_bd={bd}_kuexh={ku_ex}".format(wz=wz,ku=ku,bd=bd,ku_ex=ku_ex)
    
    else:
        print(times.at[z,'Name'])
        Vdot = pd.read_excel("C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/excel_files/Vdot_results.xlsx", sheet_name = Vdot_sheets[database])
    
        d = {}
    
        for i in range(len(Vdot)):
            d[Vdot.loc[i,"Level"] + "_sup"] = ufloat(Vdot.at[i,'Vdot_sup'], Vdot.at[i,'Vdot_sup_uncertainity'])
            d[Vdot.loc[i,"Level"] + "_exh"] = ufloat(-Vdot.at[i,'Vdot_exh'], Vdot.at[i,'Vdot_exh_uncertainity'])
        
        x = pd.DataFrame(d, index=["Vdot"])
        x = x.transpose()
        
        sz = times.at[z,'Volume flow (SZ)']
        k1 = times.at[z,'Volume flow (K1)']
        k2 = times.at[z,'Volume flow (K2)']
        ex = int(times.at[z,'Volume flow (BD)'])
        
        x.loc["BD_0_exh", "Vdot"] = ufloat(0,0)
        x.loc["BD_100_exh", "Vdot"] = ufloat(47.4,12) # data given from sven uncertainity is selected as average of uncertainities of other sensors
        
        
        
        a = max(x.at['K1_'+ str(k1) +'_sup','Vdot'] + x.at['K2_'+ str(k2) +'_sup','Vdot'], x.at['SZ_'+ str(sz) +'_exh','Vdot'] + x.at['BD_'+ str(ex) +'_exh','Vdot']) #
        
        b = max(x.at['K1_'+ str(k1) +'_exh','Vdot'] + x.at['K2_'+ str(k2) +'_exh','Vdot'] + x.at['BD_'+ str(ex) +'_exh','Vdot'], x.at['SZ_'+ str(sz) +'_sup','Vdot'] )
        
        
        Vdot_imported = (a+b)/2
        level = times.loc[z,"short name"]
        print(f"Vdot = {Vdot_imported}\n")
    #%%
    
    folder_name = times.loc[z,"short name"]
    path = "C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/results/" + folder_name
    
    # os.mkdir(path)
    
    
    
    #%% Fixed Input Data
    """
    ###############################################################################
    ##  Input Data  ###############################################################
    ###############################################################################
    """    
    experiment = z
     
    info = times.loc[[z]]
    """
    The time stamp of the initialisation of the measurement has to be setted to the 
    one where the behaviour of the measurement values for the temperature, humidity, 
    CO2-concentration and flow-speed behind the indoor hood of the ventilation 
    devices changes significantly. This time stamp has to be at the end of the 
    homogenisation period.
    """
    t0 = times["Start"][experiment]
    
    start, end = str(times["Start"][experiment] - dt.timedelta(minutes=10)), str(times["End"][experiment])
    #start, end = str(times["Start"][experiment] - dt.timedelta(minutes=10)), str(times["Start"][experiment] + dt.timedelta(hours=5))
    
    
    print(database)
    
    
    print(times["Name"][experiment])  
    print(times["short name"][experiment])  
    print("\nThe sensors excluded in this evaluation are:\n{}\n".format(times["exclude"][z]))
    """
    Document with the list of all filenames belonging to the current evaluation.
    """
    
    if "summer" in database:
        season = 'Summer'
    else:
        season = 'Winter'
        
    if "ESHL" in database:
        """
        Geometry of the indoor volume ESHL
        """
        area = (3.005 * 3.2 *2) + (4.650 * 6.417) + (1.620 + 3.817) + (1.5 * 2.4)  # #2 times l * b of bedrooms + l * b of living room from Saebu_Containermaße
        height = 2.5 # m # not considering the space after the Anhangdecke
        blocked_volume = ufloat(11.39, 1.14)
        
        
        volume = area * height
        v = volume - blocked_volume
         
    else:
        """
        Geometry of the indoor volume CBo
        """
        area = 88.7 # in m²
        height = 1.8 + 0.19 + 0.23 # m
        blocked_volume = ufloat(12.05, 1.20)
        
        volume = area * height
        v = volume - blocked_volume
    
    
    
    
    
    """
    The ventilation level choosen by the user results in a volume flow provided by
    the ventilation system.
    """
    Vdot = Vdot_imported
    
    """
    Period time of alternating ventilaton devices
    """
    T = 120 # in s
    #%%% database details
    schema = database.lower() # the schema that you work on 
    
    ''' this engine is used where ever connection is required to database'''
    engine = create_engine("mysql+pymysql://root:Password123@localhost/{}".format(schema),pool_pre_ping=True)
    
    
    
    
    #%% Main Class Tau_l
    """
    ###############################################################################
    ##  Classes  ##################################################################
    ###############################################################################
    """ 
    class Tau_l:
        end_runtime, tau_l_list, plot_log_end_list = [], [], []
        
        df_plot_list, plot_delta_list, plot_pd_SMA_list, plot_pd_SMA_end_list = [], [], [], []
        plot_log_list, df_clean_list, sec_list, plot_delta_end_list = [], [], [], []
        """
        Calculation and handover of a dataframe with the relavnt values for all the
        measurement positions in the flat. 
        1st) Indices of the dataframe are named after:
                 - measurement season (W = winter or S = summer)
                 - ventilation level (e.g. level 5 = L5)
                 - exhaust device mode (e.g. exhaust device bath L1, kitchen off = 
                   Exbk10, or exhaust device kitchen L3 = Exbk03)
                 - indoor doors open or closed (open = Do or closed = Dc)
                 - measurement position (e.g. MP1A)
                 -> example: S_L5_Exbk10_Do
        2nd) Colums: 
            'tau_lav'           = value of the local mean age of air, in h
            'sgm_tau_lav'       = simple uncertainty of tau_l,av, in h
            't0'                = timestamp of the initial averager tracer-gas 
                                    concentration, in '%Y-%m-%d %H:%M:%S'
            'dltC0_av'          = initial average offset to natural background 
                                    concentrartion of the tracer-gas of the whole 
                                    first considered ventilation cycle, in ppm
            'sgm_dltC0_av'      = simple uncertainty of dlt-C0_av, in ppm, considering: 
                                    - uncertainty natural background concentration
                                    - uncertainty sensors
                                    - standard deviation of the values belonging to
                                        to the first cycle
            'max_dltC0'          = maximum value dltC0 (the last ventilation cycle
                                before the decay curve), in ppm
            'min_dltC0'          = minimum value dltC0 (the last ventilation cycle
                                before the decay curve), in ppm        
            'te'                = timestamp of the end of the considered time
                                    evaluation intval at dlt-C0_av*0.37, 
                                    in '%Y-%m-%d %H:%M:%S'
            'dltCe_av'          = final average offset to natural background 
                                    concentrartion of the tracer-gas of the whole 
                                    last considered ventilation cycle, in ppm
            'sgm_dltCe_av'      = simple uncertainty of dlt-Ce_av, in ppm, considering: 
                                    - uncertainty natural background concentration
                                    - uncertainty sensors
                                    - standard deviation of the values belonging to
                                        to the final cycle
            'max_dltCe'         = maximum value of dltCe (values of the last 
                                considered ventilation cycle), in ppm
            'min_dltCe'         = minimum value of dltCe (values of the last 
                                considered ventilation cycle), in ppm        
            'lmb_tailav'        = tail parameter of the appoximation integral 
                                    according to DIN ISO 16000-8, in 1/h
            'sgm_lmb_tailav'    = maximum out of propagate uncertainty of the 
                                    depending parameters & standard deviation of
                                    the average logarithmic slope acc ISO 16000-8.,
                                    in 1/h
            'max_lmb_tail'      = maximum value of lmb_tail, in 1/h
            'min_lmb_tail'      = maximum value of lmb_tail, in 1/h
        """
        import pandas as pd
        import numpy as np
        # import matplotlib.pyplot as plt
        
        # CLASS OBJECT ATTRIBUTE: same for any instance of this class
        # - none jet defined -
        
        def __init__(self, filename,t0 = t0,database = database,z = z, level = level, season = season, v = v, Vdot = Vdot):
            
            # ATTRIBUTES: user defineable
            
            # Grouping attributes to "headers" indicating which type is expected
            # strings
            self.season     = season # measurement season: W = winter or S = summer
            self.level      = level # number between 1 and 9
            self.v          = v
            self.Vdot       = Vdot
            '''
                exhaust device mode (e.g. exhaust device bath L1, kitchen off = 
                   bk10, or exhaust device kitchen L3 = bk03)
            '''
            self.filename   = filename 
            '''
                Format: 
                    [room/zone: 2x ASCII-single value]_
                    MP[measurement position indicator: 2x ASCII-single value]_
                    [sensor identifier: 4x ASCII-single value]_
                    [initial timestamp: %Y%m%d_%H%M]_
                    bis_
                    [finial timestamp: %Y%m%d_%H%M].
                    [file type: xlsx or csv]
            '''
            # integers
            self.T          = 120 # Period time of alternating ventilation devices
            self.i          = 0 # row number t0_es in self.df_CO2
            self.j          = 0 # row number tend_es in self.df_CO2
            # dictionay
            self.tauloc
            
            # padas.Datetime Timestamp('%Y-%m-%d %H:%M:%S')
            self.t0         = t0
            
            # pandas.DatetimeIndex Timestamp('%Y-%m-%d %H:%M:%S')
            self.tend_es    = pd.DatetimeIndex
            
            self.database   = database.lower()
            self.start      = start
            self.end        = end
            self.df_CO2, self.sec = self.clean_sql_reg()
            self.Cout       = self.aussen() 
            
    
    
    #%%% Main method        
        # OPERATIONS/ACTIONS = Methods
        
        # Main-Method
        def tauloc(self):
            name = (self.season + '_' + self.level + '_' +
                      self.filename[0:])
            self.tau_l = pd.DataFrame(columns=(
                    'tau_lav',
                    'sgm_tau_lav',
                    't0',
                    'dltC0_av',
                    'sgm_dltC0_av',
                    'max_dltC0', 
                    'min_dltC0', 
                    'te', 
                    'dltCe_av',
                    'sgm_dltCe_av',
                    'max_dltCe', 
                    'min_dltCe',
                    'lmb_tailav',
                    'sgm_lmb_tailav',
                    'max_lmb_tail',
                    'min_lmb_tail',
                    'time_steps',
                    'row_start',
                    'row_end'),
                    index=[name])
            
            self.df_CO2     = self.co2diff()
            
            
            while not(self.t0 in self.df_CO2.index):
                self.t0 = self.t0 + dt.timedelta(seconds=1)
                print(self.t0)
            
               
            self.tau_l['t0'].loc[name] = self.t0   
            self.C0         = self.c0()
            
            self.tau_l['dltC0_av'].loc[name] = (
                self.C0['pd_SMA'].loc[self.t0])
            '''
            -----------------------------------------------------------------------
            FUTURE TASKS:
                1st) Uncertainty characteristics of 'dltC0_av'
            -----------------------------------------------------------------------
            '''    
    #        self.tau_l['sgm_dltC0_av'].loc[name] = (
    #                            self.C0['???'].loc[self.t0_es])
    #        self.tau_l['max_dltC0'].loc[name] = (
    #                            self.C0['???'].loc[self.t0_es])
    #        self.tau_l[,'min_dltC0'].loc[name] = (
    #                            self.C0['???'].loc[self.t0_es])
        
            
            self.Cend       = self.cend()
            self.i          = self.Cend[1]
            self.j          = self.Cend[2]
            self.Cend       = self.Cend[0]
            self.tend_es = self.Cend.iloc[[0]].index.tolist()[0]
            self.tau_l['te'].loc[name] = self.tend_es
            self.tau_l['dltCe_av'].loc[name] = (
                self.Cend['pd_SMA'].loc[self.tend_es])
            '''
            -----------------------------------------------------------------------
            FUTURE TASKS:
                1st) Uncertainty characteristics of 'dltC0_av'
            -----------------------------------------------------------------------
            '''    
    #        self.tau_l['sgm_dltCe_av'].loc[name] = (
    #                            self.Cend['???'].loc[self.tend_es])
    #        self.tau_l['max_dltCe'].loc[name] = (
    #                            self.Cend['???'].loc[self.tend_es])
    #        self.tau_l['min_dltCe'].loc[name] = (
    #                            self.Cend['???'].loc[self.tend_es])
            
            self.lmd_tl     = self.lmdtail()
            self.tau_l['lmb_tailav'].loc[name] = (
                1/(self.lmd_tl['lmb_tailav'].loc[self.filename[0:]]*3600)
                )
            self.tau_l['sgm_lmb_tailav'].loc[name] = (
                1/(self.lmd_tl['sgm_lmb_tailav'].loc[self.filename[0:]]*3600)
                )
            self.tau_l['max_lmb_tail'].loc[name] = (
                1/(self.lmd_tl['max_lmb_tail'].loc[self.filename[0:]]*3600)
                )
            self.tau_l['min_lmb_tail'].loc[name] = (
                1/(self.lmd_tl['min_lmb_tail'].loc[self.filename[0:]]*3600)
                )
            
         
            """
            Approch according to DIN ISO 16000-8:
                DIN ISO 2008. DIN ISO 16000-8, Innenraumluftverunreinigungen – Teil 8: 
                    Bestimmung des lokalen Alters der Luft in Gebäuden zur 
                    Charakterisierung der Lüftungsbedingungen (ISO 16000-8:2007). 
                    Berlin. Deutsches Institut für Normung e.V. and International 
                    Organization for Standardization. [equation 3]
            """
            k = 0
            sumConz = 0
            for k in range(1,len(self.df_CO2)):
                if (np.sum((self.df_CO2.iloc[[k]].index > self.C0.iloc[[0]].index) & 
                           (self.df_CO2.iloc[[k]].index < self.Cend.iloc[[0]].index))) == 1:
                    sumConz = sumConz + self.df_CO2['pd_SMA'].iloc[k]
                    if np.sum(self.df_CO2.iloc[[k]].index >= self.Cend.iloc[[0]].index) == 1:
                        break
        
            tau_lav = (
                    1/(3600*self.C0['Delta_in-out_C'].iloc[0])*(
                            self.sec * ( self.C0['Delta_in-out_C'].iloc[0]/2 + 
                                   sumConz +
                                   self.Cend['pd_SMA'].iloc[0]/2 
                                   ) + (
                                           self.Cend['pd_SMA'].iloc[0]/
                                           self.lmd_tl['lmb_tailav'].iloc[0]
                                           )
                                   )
                            )
            self.tau_l.loc[name, 'tau_lav'] = tau_lav
            
            self.tau_l.loc[name, 'time_steps'] = self.sec
            self.tau_l.loc[name, 'row_start'] = self.i 
            self.tau_l.loc[name, 'row_end'] = self.j         
            
            #######################################################################
            
            self.plot_delta = self.co2diff()[['runtime','Delta_in-out_C']]
            new_col = [self.filename if col != "runtime" else str(col) for col in self.plot_delta.columns]
            self.plot_delta.columns = new_col
            self.plot_delta_list.append(self.plot_delta)
            self.plot_delta_end_list.append(self.Cend["Delta_in-out_C"].iloc[0])
            self.end_runtime.append(self.Cend["runtime"].iloc[0])
            
            
            self.plot_pd_SMA = self.co2diff()[['runtime','pd_SMA']]
            new_col = [self.filename if col != "runtime" else str(col) for col in self.plot_pd_SMA.columns]
            self.plot_pd_SMA.columns = new_col
            self.plot_pd_SMA_list.append(self.plot_pd_SMA)
            self.plot_pd_SMA_end_list.append(self.Cend["pd_SMA"].iloc[0])
            
            
            self.plot_log = self.co2diff()[['runtime', 'log_value']]
            new_col = [ self.filename if col != "runtime" else str(col) for col in self.plot_log.columns]
            self.plot_log.columns = new_col
            self.plot_log_list.append(self.plot_log)        
            self.plot_log_end_list.append(self.Cend["log_value"].iloc[0])
            #######################################################################
            
            prRed(self.filename)
            return self.tau_l
    #%%% Sub Methods    
       # Sub-methods
    
        def clean_sql_reg(self):
            self.sensor_name = self.filename
            
            accuracy1 = 50 # it comes from the equation of uncertainity for testo 450 XL
            accuracy2 = 0.02 # ±(50 ppm CO2 ±2% of mv)(0 to 5000 ppm CO2 )
            
            accuracy3 = 50 # the same equation for second testo 450 XL
            accuracy4 = 0.02
            
            accuracy5 = 75 # # the same equation for second testo 480
            accuracy6 = 0.03 # Citavi Title: Testo AG
            
            '''
            The following if esle statement is writtten to import the right data 
            for calibration offset equation
            There are two time periods where calibration was done and this
            '''
            
            if (self.database == "cbo_summer") or (self.database == "cbo_winter") or (self.database == "eshl_winter"):
                engine = create_engine("mysql+pymysql://root:Password123@localhost/{}".format("cbo_calibration"),pool_pre_ping=True)
        #        engine = create_engine("mysql+pymysql://root:@34.107.104.23/{}".format("cbo_calibration"),pool_pre_ping=True)
        
            elif self.database == "eshl_summer":
                engine = create_engine("mysql+pymysql://root:Password123@localhost/{}".format("eshl_calibration"),pool_pre_ping=True)
        #        engine = create_engine("mysql+pymysql://root:@34.107.104.23/{}".format("eshl_calibration"),pool_pre_ping=True)
        
            else:
                print("Please select a correct database")
            
            '''standard syntax to import sql data as dataframe
            engine 1 is measurement campagin experimentl data and engine is calibration data'''
            engine1 = create_engine("mysql+pymysql://root:Password123@localhost/{}".format(self.database),pool_pre_ping=True)
            '''Calibration data is imported '''
            reg_result = pd.read_sql_table("reg_result", con = engine).drop("index", axis = 1)
            '''Calibration data for the particular sensor alone is filtered '''
            res = reg_result[reg_result['sensor'].str.lower() == self.sensor_name].reset_index(drop = True)
            
            '''This is to filter the HOBOs from testos, The hobos will have a res variable Testos will not have
            because they dont have experimantal calibration offset'''
            if res.shape[0] == 1:
                ''' The imported sql data is cleaned and columns are renamed to suit to out calculation'''
                self.sensor_df = pd.read_sql_query("SELECT * FROM {}.{} WHERE datetime BETWEEN '{}' AND '{}'".format(self.database, self.sensor_name, self.start, self.end) , engine1).drop('index', axis =1)
                self.sensor_df['CO2_ppm_reg'] = self.sensor_df.eval(res.loc[0, "equation"])    
                self.sensor_df = self.sensor_df.rename(columns = {'CO2_ppm':'CO2_ppm_original', 'CO2_ppm_reg': 'C_CO2 in ppm'})
                self.sensor_df = self.sensor_df.drop_duplicates(subset=['datetime'])
                self.sensor_df = self.sensor_df.loc[:, ["datetime", "C_CO2 in ppm", "CO2_ppm_original"]]
                self.sensor_df = self.sensor_df.dropna()
                '''This is the absolute uncertainity at each point of measurement saved in the
                dataframe at each timestamp Ref: equation D2 in DIN ISO 16000-8:2008-12'''
                
                
                '''For ESHL summer ideally we take mean of all three sensors and also propogate 
                the uncertainities of al three testo sensors, This is not done here at the moment
                But to get the most uncertainity possible we peopogte the uncertainity first'''
                # Why RSE ? https://stats.stackexchange.com/questions/204238/why-divide-rss-by-n-2-to-get-rse
                self.sensor_df["s_meas"] =  np.sqrt(np.square((self.sensor_df["C_CO2 in ppm"] * accuracy2)) + np.square(accuracy1) + np.square((self.sensor_df["C_CO2 in ppm"] * accuracy4)) + np.square(accuracy3) + np.square((self.sensor_df["C_CO2 in ppm"] * accuracy6)) + np.square(accuracy5)+ np.square(res.loc[0, "rse"])) 
                # Die Messunsicherheit hängt sicher in einem bestimmten Umfang vom Konzentrationsbereich ab.DIN ISO 16000-8:2008-12 (page 36)
        
                x = self.sensor_df["datetime"][2] - self.sensor_df["datetime"][1]
                self.sec = int(x.total_seconds())
                
                """
                Creating a runtime column with t0 as 0 or centre of the time axes
                """
                t0_cd = self.sensor_df['datetime'].loc[0]
                
                while not(self.t0 in self.sensor_df["datetime"].to_list()):
                    self.t0 = self.t0 + dt.timedelta(seconds=1)
                    print(self.t0)
                    
                dtl_t0 = (self.t0 - t0_cd)//dt.timedelta(seconds=1)
                
                """
                Calucates the elapsed time stored in the array x as an interger of seconds
                """
                endpoint = len(self.sensor_df) * self.sec - dtl_t0
                
                """
                Creates an array starting with 0 till endpoint with stepsize sec.
                """
                x = np.arange(-dtl_t0,endpoint,self.sec)
                
                self.sensor_df['runtime'] = x
                
                self.sensor_df2 = self.sensor_df.set_index('datetime')
                self.rhg = pd.date_range(self.sensor_df2.index[0], self.sensor_df2.index[-1], freq=str(self.sec)+'S')   
                self.sensor_df = self.sensor_df2.reindex(self.rhg).interpolate()
                
                return self.sensor_df, self.sec
        
            else:
                '''The steps taken place before simply repeat here below for testo sensors
                '''
                self.sensor_df = pd.read_sql_query("SELECT * FROM {}.{} WHERE datetime BETWEEN '{}' AND '{}'".format(self.database, self.sensor_name, self.start, self.end) , engine) # selects a table from a database according to table name
                self.sensor_df = self.sensor_df.drop_duplicates(subset=['datetime'])
                self.sensor_df = self.sensor_df.loc[:, ["datetime", "CO2_ppm"]]
                self.sensor_df = self.sensor_df.rename(columns = {"CO2_ppm":"C_CO2 in ppm"})
                self.sensor_df["s_meas"] =  np.sqrt(np.square((self.sensor_df["C_CO2 in ppm"] * accuracy2)) + np.square(accuracy1) + np.square((self.sensor_df["C_CO2 in ppm"] * accuracy4)) + np.square(accuracy3) + np.square((self.sensor_df["C_CO2 in ppm"] * accuracy6)) + np.square(accuracy5))   
                self.sensor_df = self.sensor_df.dropna()
                x = self.sensor_df["datetime"][2] - self.sensor_df["datetime"][1]
                self.sec = int(x.total_seconds())
                
                """
                Creating a runtime column with t0 as 0 or centre of the time axes
                """
                t0_cd = self.sensor_df['datetime'].loc[0]
                
                
                while not(self.t0 in self.sensor_df["datetime"].to_list()):
                    self.t0 = self.t0 + dt.timedelta(seconds=1)
                    print(self.t0)
                
                dtl_t0 = (self.t0 - t0_cd)//dt.timedelta(seconds=1)
                """
                Calucates the elapsed time stored in the array x as an interger of seconds
                """
                endpoint = len(self.sensor_df) * self.sec - dtl_t0
                
                """
                Creates an array starting with 0 till endpoint with stepsize sec.
                """
                x = np.arange(-dtl_t0,endpoint,self.sec)
                
                self.sensor_df['runtime'] = x
                
                self.sensor_df2 = self.sensor_df.set_index('datetime')
                self.rhg = pd.date_range(self.sensor_df2.index[0], self.sensor_df2.index[-1], freq=str(self.sec)+'S')   
                self.sensor_df = self.sensor_df2.reindex(self.rhg).interpolate()
                    
                return self.sensor_df, self.sec
        
        
        def aussen(self):
            if self.database == "cbo_summer":
                self.Cout = {'meanCO2': 445.1524174626867,
                        'sgm_CO2': 113.06109664245112,
                        'maxCO2': 514.3716999999999,
                        'minCO2': 373.21639999999996}
                self.cout_mean, self.cout_max, self.cout_min = 445.1524174626867, 514.3716999999999, 373.21639999999996
                return self.Cout
            else:
                
                accuracy1 = 50 # it comes from the equation of uncertainity for testo 450 XL
                accuracy2 = 0.02 # ±(50 ppm CO2 ±2% of mv)(0 to 5000 ppm CO2 )
                
                accuracy3 = 50 # the same equation for second testo 450 XL
                accuracy4 = 0.02
                
                accuracy5 = 75 # # the same equation for second testo 480
                accuracy6 = 0.03 # Citavi Title: Testo AG
                
                '''
                The following if esle statement is writtten to import the right data 
                for calibration offset equation
                There are two time periods where calibration was done and this
                '''
                self.database = self.database.lower()
                if (self.database == "cbo_summer") or (self.database == "cbo_winter") or (self.database == "eshl_winter"):
                    engine = create_engine("mysql+pymysql://root:Password123@localhost/{}".format("cbo_calibration"),pool_pre_ping=True)
            #        engine = create_engine("mysql+pymysql://root:@34.107.104.23/{}".format("cbo_calibration"),pool_pre_ping=True)
            
                elif self.database == "eshl_summer":
                    engine = create_engine("mysql+pymysql://root:Password123@localhost/{}".format("eshl_calibration"),pool_pre_ping=True)
            #        engine = create_engine("mysql+pymysql://root:@34.107.104.23/{}".format("eshl_calibration"),pool_pre_ping=True)
            
                else:
                    print("Please select a correct database")
                
                '''standard syntax to import sql data as dataframe
                engine 1 is measurement campagin experimentl data and engine is calibration data'''
                engine1 = create_engine("mysql+pymysql://root:Password123@localhost/{}".format(self.database),pool_pre_ping=True)
                '''Calibration data is imported '''
                reg_result = pd.read_sql_table("reg_result", con = engine).drop("index", axis = 1)
                '''Calibration data for the particular sensor alone is filtered '''
                res = reg_result[reg_result['sensor'].str.lower() == "außen"].reset_index(drop = True)
                
                '''This is to filter the HOBOs from testos, The hobos will have a res variable Testos will not have
                because they dont have experimantal calibration offset'''
                if res.shape[0] == 1:
                    ''' The imported sql data is cleaned and columns are renamed to suit to out calculation'''
                    self.sensor_df3 = pd.read_sql_query("SELECT * FROM {}.{} WHERE datetime BETWEEN '{}' AND '{}'".format(self.database, "außen", self.start, self.end) , engine1).drop('index', axis =1)
                    self.sensor_df3['CO2_ppm_reg'] = self.sensor_df3.eval(res.loc[0, "equation"])    
                    self.sensor_df3 = self.sensor_df3.rename(columns = {'CO2_ppm':'CO2_ppm_original', 'CO2_ppm_reg': 'C_CO2 in ppm'})
                    self.sensor_df3 = self.sensor_df3.drop_duplicates(subset=['datetime'])
                    self.sensor_df3 = self.sensor_df3.loc[:, ["datetime", "C_CO2 in ppm", "CO2_ppm_original"]]
                    self.sensor_df3 = self.sensor_df3.dropna()
                    '''This is the absolute uncertainity at each point of measurement saved in the
                    dataframe at each timestamp Ref: equation D2 in DIN ISO 16000-8:2008-12'''
                    
                    
                    '''For ESHL summer ideally we take mean of all three sensors and also propogate 
                    the uncertainities of al three testo sensors, This is not done here at the moment
                    But to get the most uncertainity possible we peopogte the uncertainity first'''
                    # Why RSE ? https://stats.stackexchange.com/questions/204238/why-divide-rss-by-n-2-to-get-rse
                    self.sensor_df3["s_meas"] =  np.sqrt(np.square((self.sensor_df3["C_CO2 in ppm"] * accuracy2)) + np.square(accuracy1) + np.square((self.sensor_df3["C_CO2 in ppm"] * accuracy4)) + np.square(accuracy3) + np.square((self.sensor_df3["C_CO2 in ppm"] * accuracy6)) + np.square(accuracy5)+ np.square(res.loc[0, "rse"])) 
                    # Die Messunsicherheit hängt sicher in einem bestimmten Umfang vom Konzentrationsbereich ab.DIN ISO 16000-8:2008-12 (page 36)
            
                    x = self.sensor_df3["datetime"][2] - self.sensor_df3["datetime"][1]
                    self.sec3 = int(x.total_seconds())
                    
                    """
                    Creating a runtime column with t0 as 0 or centre of the time axes
                    """
                    t0_cd = self.sensor_df3['datetime'].loc[0]
                    
                    while not(self.t0 in self.sensor_df3["datetime"].to_list()):
                        self.t0 = self.t0 + dt.timedelta(seconds=1)
                        print(self.t0)
                        
                    dtl_t0 = (self.t0 - t0_cd)//dt.timedelta(seconds=1)
                    
                    """
                    Calucates the elapsed time stored in the array x as an interger of seconds
                    """
                    endpoint = len(self.sensor_df3) * self.sec3 - dtl_t0
                    
                    """
                    Creates an array starting with 0 till endpoint with stepsize sec3.
                    """
                    x = np.arange(-dtl_t0,endpoint,self.sec3)
                    
                    self.sensor_df3['runtime'] = x
                    
                    self.sensor_df2 = self.sensor_df3.set_index('datetime')
                    self.rhg = pd.date_range(self.sensor_df2.index[0], self.sensor_df2.index[-1], freq=str(self.sec3)+'S')   
                    self.au_mean = self.sensor_df2.reindex(self.rhg).interpolate()
                    
                    self.au_mean['C_CO2 in ppm_out'] = self.au_mean['C_CO2 in ppm']
                    self.cout_max = self.au_mean['C_CO2 in ppm_out'].max()
                    self.cout_min = self.au_mean['C_CO2 in ppm_out'].min()
                    self.cout_mean = self.au_mean['C_CO2 in ppm_out'].mean()
                    
                    """
                    The default value (499±97)ppm (kp=2) has been calculated as the average CO2-
                    concentration of the available outdoor measurement data in 
                    ...\CO2-concentration_outdoor\.
                    However the value should be setted as a list of datapoints for the natural
                    outdoor concentration for a time inverval covering the measurement interval.
                    
                    In future it would be great to have a dataframe with CO2-concentrations for 
                    coresponding time stamps.
                    """
                    self.Cout = {'meanCO2':self.cout_mean, 
                            'sgm_CO2':self.au_mean["s_meas"].mean(), # More clarification needed on uncertainity
                            'maxCO2':self.cout_max,
                            'minCO2':self.cout_min}
                    return self.Cout
        """
        If nothing else is passed this calculates the difference of the
        measured CO₂-concentration to the averager outdoor CO₂-concentration.
        Where the CO₂-concentration is by default setted to (499±97)ppm (kp=1).
        Thise value has been calculated as the average CO₂-concentration of the
        available outdoor measurement data in ...\CO2-concentration_outdoor\.
    
        """
    
    
        def co2diff(self):
            import numpy as np
            # Calculates difference of measured to background/outdoor cconcentration
            '''
            -----------------------------------------------------------------------
            FUTURE TASKS:
                1st) Check whether a data frame of outdoor CO₂-concentration for  
                    several time stamps within the considered time interval has  
                    been passed. If "True" than use those concentrations instead.
                2nd) Calculate the uncertainty- and range-measures of the 
                    'Delta_in-out_C'-values.
            -----------------------------------------------------------------------
            '''
    #        self.df_CO2['Delta_in-out_C'] = (self.df_CO2['C_CO2 in ppm'] - 
    #                   self.Cout['meanCO2'])
            
            self.df_CO2['Delta_in-out_C']=self.df_CO2['C_CO2 in ppm'].subtract(self.Cout['meanCO2'])
            
            # Moving average of measured data covering a time interval of T
            '''
            -----------------------------------------------------------------------
            FUTURE TASKS:
                1st) Calculate the uncertainty- (standard deviation) and range-
                    measures of the 'pd_SMA'-values over the time interval of T
            -----------------------------------------------------------------------
            '''
            self.df_CO2['pd_SMA'] = self.df_CO2['Delta_in-out_C'].rolling(
                    window=((self.T//self.sec)+1)
                    ).mean()
            
            # Natural logarithm of the moving average
            '''
            -----------------------------------------------------------------------
            FUTURE TASKS:
                1st) Calculate the uncertainty- and range-
                    measures of the 'log_value'-values over the time interval of T
            -----------------------------------------------------------------------
            '''
            self.df_CO2['log_value'] = np.log(self.df_CO2['pd_SMA'])
            
            return self.df_CO2
       
        """
        Calculates the initial concentration of the decay curve as the average-value
        of the last ventilation cycle (T=120 s) of the ventilation devices.
        The value is saved together with the last equally spaced timestamp t0_es and
        runtime value.
        ---------------------------------------------------------------------------
        FUTURE TASKS:
            1st) Include a check whether the data frame df_CO2 has already been
                supplemented by the column 'Delta_in-out_C'. In case of "False" 
                co2diff(self) should be performed.
        ---------------------------------------------------------------------------
        """
        
        def c0(self):
            self.df_C0 = self.df_CO2.truncate(
                    before=(pd.to_datetime(self.t0,format = '%Y-%m-%d %H:%M:%S') - 
                            pd.to_timedelta(120, unit='s')), 
                    after=self.t0)
            self.df_C0.loc[self.t0,'C_CO2 in ppm'] = self.df_C0['C_CO2 in ppm'].mean()
    
            self.C0 = self.df_C0.loc[[self.t0]]
            
            return self.C0
            
    
    
               
        """
        Defines the final concentration of the concentration decay curve Cend_2 as 
        the value which represents 37 % of the total decay potential above the
        outdoor concentration. This is because at this point the error of the 
        integral will be minimized.
        1) Maas, A. Prof. Dr.-Ing. (1995). Experimentelle Quantifizierung des 
            Luftwechsels bei Fensterlüftung, Dissertation, Universität 
            Gesamthochschule Kassel. Kassel, 106pp., page 57.
        2) Roulet, C.-A. and Vandaele, L. (1991). Air flow patterns within 
            buildings, measurement techniques: Air leakage measurement methods, Air 
            flow measurement methods, Measurement methods related to efficiency, 
            Measurements on ventilation systems. Coventry:, pages III.21 - III.22.
        3) Sherman, M.H. (1990), Tracer-gas techniques for measuring ventilation in 
            a single zone. Building and Environment, 25(4), pages 365–374.
        ---------------------------------------------------------------------------
        FUTURE TASKS:
            1st) In case a time series of outdoor concentrations has already been 
                implemented and passed to the current instance of Tau_l 
                self.Cout['meanCO2'] and its uncertainty measures have to be
                calculated first.
            2nd) Uncluding the calculation of the uncertainty measures of self.Cend.
        ---------------------------------------------------------------------------
        """
        def cend(self):
            self.Cend37 = round(
                    (self.C0['C_CO2 in ppm'].iloc[0] - self.Cout['meanCO2'])*0.37,
                    2)
            
            """
            Extracting the first timestamp and CO2-concentration where the moving
            average of a whole ventilation cycle of the decay-curve droped below the 
            requiered minimal error concentration.
            """    
    
            self.a = self.df_CO2.loc[self.df_CO2['runtime'].le(0)]
            self.i = len(self.a) - 1
            
            self.b = self.df_CO2.loc[self.df_CO2['pd_SMA'].le(self.Cend37)]
            
            if self.b.empty:
                self.j = len(self.df_CO2) - 1
                self.Cend = self.df_CO2.iloc[[self.j]]
                self.Cend_value = round(self.Cend.iloc[0].loc["pd_SMA"], 2)
                self.Cend_percent = round((self.Cend_value * 100)/(self.C0['C_CO2 in ppm'].iloc[0] - self.Cout['meanCO2']) , 2)
                print(f"{self.filename} has not reached {self.Cend37} ppm i.e 37% of its initial concentration, But reached {self.Cend_percent}% i.e {self.Cend_value} ppm")
            else:
                self.j = int(self.b['runtime'][0]/self.sec + self.i)
                self.Cend = self.b[:1]        
            self.Cend.loc[:,['log_value']].iat[0,0] = np.log(float(self.Cend['pd_SMA']))
                 
            return self.Cend, self.i, self.j 
        
       
        """
        Calculate slope of the tail
        ---------------------------
        1st) Creating an empty dataframe indexed from t0_es till tend_es-10s with
            equally spaced timestamps of 10 seconds.
        2nd) Fill the dataframe
        """
        def lmdtail(self):
            if (self.i, self.j) == ( 0, 0):
                self.Cend       = self.cend()
                self.i          = self.Cend[1]
                self.j          = self.Cend[2]
            
            # 1st)
            self.df_lmd_tl = pd.date_range(start=self.t0, 
                                           freq=str(self.sec)+'S', 
                                           periods= (self.j-self.i))
            self.df_lmd_tl = pd.DataFrame(index=self.df_lmd_tl, 
                                        columns=['t - t_e', 
                                                 'ln(C_te / C_t)', 
                                                 'lmbd_tail'])
            # 2nd)
            for k in range(self.i,self.j):
                self.df_lmd_tl.loc[self.df_CO2.iloc[[k]].index.tolist(), 't - t_e'] = (
                        self.df_CO2['runtime'].iloc[k] - 
                        self.Cend['runtime'].iloc[0])
    #            self.df_lmd_tl['ln(C_te / C_t)'].loc[self.df_CO2.iloc[[k]].index.tolist()] = (
    #                     self.Cend['log_value'].iloc[0] - 
    #                     self.df_CO2['log_value'].iloc[k])
                self.df_lmd_tl.loc[self.df_CO2.iloc[[k]].index.tolist(),'ln(C_te / C_t)'] = (
                         self.Cend['log_value'].iat[0] - 
                         self.df_CO2['log_value'].iat[k])            
        
            self.df_lmd_tl['lmbd_tail'] = (
                        self.df_lmd_tl['ln(C_te / C_t)'] / 
                        self.df_lmd_tl['t - t_e'])
            
            self.lmd_tl = pd.DataFrame(columns=('lmb_tailav', 
                                              'sgm_lmb_tailav', 
                                              'max_lmb_tail',
                                              'min_lmb_tail'))
            self.lmd_tl.loc[self.filename[0:],'lmb_tailav'] = np.mean(
                    self.df_lmd_tl['lmbd_tail'])
            
            self.lmd_tl.loc[self.filename[0:],'sgm_lmb_tailav'] = (
                    np.std(self.df_lmd_tl['lmbd_tail']))
            
            self.lmd_tl.loc[self.filename[0:],'max_lmb_tail'] = np.max(
                    self.df_lmd_tl['lmbd_tail'])
            
            self.lmd_tl.loc[self.filename[0:],'min_lmb_tail'] = np.min(
                    self.df_lmd_tl['lmbd_tail'])
            
            return self.lmd_tl
        
        # Plot-Methods
        """
        Plotting the logarithmic decay-curves including the selected initial and
        final timestamps in order check whether the right time interval has been
        chosen for the evaluation.
        ---------------------------------------------------------------------------
        FUTURE TASKS:
            1st) In order to save computation time the first methods should just
                be performed in case the necessary values are empty or not existing.
        ---------------------------------------------------------------------------
        """
        def pltLogData(self):
            
            self.df_CO2     = self.co2diff()
            self.C0         = self.c0()
            self.Cend       = self.cend()
            
            ax = plt.gca()
            plt.figure()
    
            plt.title(self.filename[0:])
            self.df_CO2.plot(x='runtime', 
                             y='log_value', 
                             ax=ax)
            self.C0.plot(x='runtime',
                         y='log_value', 
                         marker='o', color='red', 
                         ax=ax)
            self.Cend[0].plot(x='runtime', 
                           y='log_value', 
                           marker='o', 
                           color='red', 
                           ax=ax)
            return
            
        """
        Ploting the delta concentration relative to the outdoor concentration of 
        each timestamp in comparission with the moving average delta concentration
        of whole ventilation cycle.
        ---------------------------------------------------------------------------
        FUTURE TASKS:
            1st) In order to save computation time the first methods should just
                be performed in case the necessary values are empty or not existing.
        ---------------------------------------------------------------------------
        """
        def pltDeltaMoving(self):
            
            self.df_CO2     = self.co2diff()
            
            ax = plt.gca()
            plt.figure()
            plt.title(self.filename[0:])
            self.df_CO2.plot(x='runtime', y='Delta_in-out_C',ax=ax)
            self.df_CO2.plot(x='runtime', y='pd_SMA',ax=ax)
            
            return
        
        def uncertainity(self, s_c0 = 0.05): # contcat hoesen@vdi.de 
            self.tau_l = self.tauloc()
            self.s_meas = self.df_CO2.truncate(before = self.C0.index[0], after=self.Cend.index[0])
            # Comment about the origin 
            self.ns_meas = self.s_meas['s_meas'].sum()
            self.n = len(self.s_meas)
            self.sa_num = np.sqrt(self.ns_meas) * (self.sec) * ((self.n - 1)/self.n) # Taken from DIN ISO 16000-8:2008-12, Equation D2 units are cm3.m-3.sec
            self.a_tot = 3600 * self.tau_l['tau_lav'][0] * self.c0()['C_CO2 in ppm'][0] 
            self.a_rest = 3600 * self.tau_l['tau_lav'][0] * self.Cend['C_CO2 in ppm'][0]
            
            self.s_lambda_df = self.df_CO2['log_value'].truncate(before = str(self.C0.index[0]), after  = str(self.Cend.index[0]))
            self.std_s_lambda_df = self.s_lambda_df.std()
            self.mean_s_lambda_df = self.s_lambda_df.mean()
            self.phi_e = self.s_lambda_df.iloc[-1]
            
            self.s_lambda = self.df_lmd_tl["lmbd_tail"].std()/self.df_lmd_tl["lmbd_tail"].mean() #The relative standard deviation sλ of the correlation coefficient (−λ)

            self.s_phi_e = self.std_s_lambda_df/self.phi_e
            self.s_rest = np.sqrt(pow(self.s_lambda,2) + pow(self.s_phi_e,2))
            
            
            # self.s_rest = 0.023 #2.3% taken as a temporary value, to be calculated
            self.sa_rest = self.s_rest * self.a_rest
            
            self.s_area = np.sqrt(pow(self.sa_num,2) + pow(self.sa_rest,2))/self.a_tot
            
            self.s_c0 = s_c0 # 5 percent initial concentration difference
            self.s_total = np.sqrt(pow(self.s_area,2) + pow(self.s_c0,2))
    
            
            self.tau_l["sgm_tau_lav"][0] = self.tau_l["tau_lav"][0] * self.s_total
    #        self.tau_l["tau_lav_uncertainity"] = ufloat(self.tau_l.iat[0,0], self.tau_l.iat[0,1])
            self.tau_l.insert(loc = 0, column = "tau_lav_uncertainity", value = ufloat(self.tau_l.iat[0,0], self.tau_l.iat[0,1]) ,allow_duplicates = True )
            self.tau_l_list.append(self.tau_l)
                    
            return self.s_total, self.s_area, self.s_rest, self.s_phi_e, self.s_lambda
        
        def result(self):
            
            self.df_eps = pd.concat(self.tau_l_list)
            """
            nominal air-exchange rate τnom
            """
            # tau_nom = v/Vdot['Vdot_av'] # hour
            
            self.tau_nom = self.v/self.Vdot # hour    
                
            """
            spacial average of air age 〈τ̅〉
            -------------------------------------------------------------------------------
            FUTURE TASK:
                1st) Uncertainty measures including error propagation and standard
                    deviation of the mean value
                2nd) Minimum and maximum measured age of air
            -------------------------------------------------------------------------------
            """
            self.tau_av = self.df_eps['tau_lav_uncertainity'].sum()/self.df_eps['tau_lav_uncertainity'].count()    
            
            """
            local air exchange indicator εN,j
            SOURCE:
                Skaaret, Eimund, 1986, Contaminant removal performance in terms of 
                ventilation effectiveness., Environment International 12 (1-4), pp. 
                419–427. DOI: 10.1016/0160-4120(86)90057-7. page 421
            -------------------------------------------------------------------------------
            FUTURE TASK:
                1st) Uncertainty measures including error propagation
            -------------------------------------------------------------------------------
            """
              
            self.df_eps.insert(loc = 0, column = 'eps_aNj', value = self.tau_av/self.df_eps['tau_lav_uncertainity'],allow_duplicates = True )   
            self.df_eps.insert(loc = 0, column = 'eps_aNj_n',value = unumpy.nominal_values(self.tau_av/self.df_eps['tau_lav_uncertainity']),allow_duplicates = True)
            
            """
            local air exchange index εai
            SOURCE:
                E. Mundt; H. M. Mathisen; P. Nielsen; A. Moser, 2004, Ventilation 
                Effectiveness., Brussels: Rehva (REHVA Guidebook, 2), 
                ISBN: 978-1-52-311574-7. page 24
            -------------------------------------------------------------------------------
            FUTURE TASK:
                1st) Uncertainty measures including error propagation
            -------------------------------------------------------------------------------
            """
            self.df_eps.insert(loc = 0, column = 'eps_ai',value = self.tau_nom/self.df_eps['tau_lav_uncertainity'],allow_duplicates = True)
            self.df_eps.insert(loc = 0, column = 'eps_ai_n',value = unumpy.nominal_values(self.tau_nom/self.df_eps['tau_lav_uncertainity']),allow_duplicates = True)

            self.ea = self.tau_nom/(2 * self.tau_av)
            
            return self.df_eps, self.ea, self.tau_av
             
        def print_results(self):
            print("Short Name of the experiment = {}".format(self.level))
            print("season of the experiment = {}".format(self.database))
            
            print("Volume flow Vdot = {:0.2f} m3/hr".format(self.Vdot))
            print("\nAverage air age is: {:0.2f} hours".format(self.tau_av))
            print("\nGlobal air exchange efficiency is: {:0.2f}".format(self.ea))
            
        def plot_test_2(self, path):
            runtime_end = max(self.end_runtime) + 1000
            plt.figure()
            ax1 = plt.gca()

            
            plt.title("Abklingkurve all Delta_in-out_C")    
            for i in self.plot_delta_list: # all raw data plotted in the background in silver
                i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax1, color = 'silver')
            #ax1.get_legend().remove()   
            for i, j, k in list(zip(self.plot_pd_SMA_list, self.end_runtime, self.plot_pd_SMA_end_list)) :
                if "1L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax1, color = '#179C7D')
                    ax1.scatter(j, k, marker='o', color = '#179C7D')
                elif "2L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax1, color = '#F29400')
                    ax1.scatter(j, k, marker='o', color = '#F29400')
                elif "3L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax1, color = '#1F82C0')
                    ax1.scatter(j, k, marker='o', color = '#1F82C0')
                elif "4L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax1, color = '#E2001A')
                    ax1.scatter(j, k, marker='o', color = '#E2001A')
                elif "1" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax1, color = '#179C7D')
                    ax1.scatter(j, k, marker='o', color = '#179C7D')
                elif "2" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax1, color = '#F29400')
                    ax1.scatter(j, k, marker='o', color = '#F29400')
                elif "3" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax1, color = '#1F82C0')
                    ax1.scatter(j, k, marker='o', color = '#1F82C0')
                elif "4" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax1, color = '#E2001A')
                    ax1.scatter(j, k, marker='o', color = '#E2001A')
                      
            
            plt.ylim(top = 3000)
            
            # =============================================================================
            # The Part below is still a mystery , needs more understanding.
            # =============================================================================
            #handles, labels = plt.gca().get_legend_handles_labels()
            #unique = [(h,l) for i, (h, l) in enumerate(zip(handles, labels)) if "Delta" not in l]
            #unique.sort(key = operator.itemgetter(1))
            #ax1.legend(*zip(*unique)) 
            
            
            handles, labels = plt.gca().get_legend_handles_labels()
            unique = [(l,h) for i, (h, l) in enumerate(zip(handles, labels)) if "Delta" not in l]
            unique.sort(key = operator.itemgetter(0))
            d = dict(unique)
            
            unique1 = {key: value for key, value in d.items() if "1" in key}
            leg1 = ax1.legend(unique1.values(), unique1.keys(),  loc='upper right', bbox_to_anchor=(0.79, 1), title = "$\mathbf{Child\/1}$")
            
            unique2 = {key: value for key, value in d.items() if "2" in key}
            leg2 = ax1.legend(unique2.values(), unique2.keys(),  loc='upper right', bbox_to_anchor=(0.86, 1), title = "$\mathbf{Child\/2}$")
            
            unique3 = {key: value for key, value in d.items() if "3" in key}
            leg3 = ax1.legend(unique3.values(), unique3.keys(), loc='upper right', bbox_to_anchor=(0.93, 1), title = "$\mathbf{Bedroom}$")
            
            unique4 = {key: value for key, value in d.items() if "4" in key}
            leg4 = ax1.legend(unique4.values(), unique4.keys(), loc='upper right', bbox_to_anchor=(1, 1), title = "$\mathbf{Bathroom}$")
            
            ax1.add_artist(leg1)
            ax1.add_artist(leg2)
            ax1.add_artist(leg3)
            
            plt.axvline(x = 0, c = 'k', ls = '--',lw = 0.5, label='Start of the Experiment')
    #        plt.axvline(x = 7440, c = 'k',ls = '--',lw = 0.5, label='End of the Experiment')
            plt.text(x = 10, y = -250, s = "Start of evaluation " , c = 'k')
    #        plt.text(x = 7470, y = -250, s = "End of evaluation", c = 'k') 
            plt.xlabel("runtime t in s")
            plt.ylabel(r'$\mathregular{\Delta  C_{CO_2}(t, }$' + r"$\bf{" + 'x' + "}$" + ')')
            plt.show()
    
    ###############################################################################
            c = [(23,125,156),(242,0,148),(31,192,130),(226,26,0),(177,0,200)] # color codes given by Fraunhofer 
            c = ['#179C7D','#F29400','#1F82C0','#E2001A','#B1C800'] # hex values of Fraunhofer colour codes obtained via internet
            
            
            f, (ax1, ax2) = plt.subplots(2, 1, sharex=True) # initialize a empty figure with 2 subplots
            
            #for i in plot_pd_SMA_list:
            #    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax1)
            for i, j, k in list(zip(self.plot_pd_SMA_list, self.end_runtime, self.plot_pd_SMA_end_list)) : # a loop to find out sensors belonging to each room and color them with same color and line style, 4 rooms so 4 elif statements
                if "1L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax1, color = '#179C7D')
                    ax1.scatter(j, k, marker='o', color = '#179C7D')
                elif "2L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax1, color = '#F29400')
                    ax1.scatter(j, k, marker='o', color = '#F29400')
                elif "3L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax1, color = '#1F82C0')
                    ax1.scatter(j, k, marker='o', color = '#1F82C0')
                elif "4L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax1, color = '#E2001A')
                    ax1.scatter(j, k, marker='o', color = '#E2001A')
                elif "1" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax1, color = '#179C7D')
                    ax1.scatter(j, k, marker='o', color = '#179C7D')
                elif "2" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax1, color = '#F29400')
                    ax1.scatter(j, k, marker='o', color = '#F29400')
                elif "3" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax1, color = '#1F82C0')
                    ax1.scatter(j, k, marker='o', color = '#1F82C0')
                elif "4" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax1, color = '#E2001A')
                    ax1.scatter(j, k, marker='o', color = '#E2001A')
                     
            ax1.set_ylabel(r'$\mathregular{\Delta  C_{CO_2}(t, }$' + r"$\bf{" + 'x' + "}$" + ')') # had to use set_label to set mathematical expression as label
            handles, labels = ax1.get_legend_handles_labels() # this line gets the values to handles (matplotlib object) and lables in the same order as in the graph
            unique = [(l, h) for i, (h, l) in enumerate(zip(handles, labels))] # Zip function makes two values into a list of tuples these tuples are later sorted
            unique.sort(key = operator.itemgetter(0)) # these tuples are sorted and later converted into a dictinory to pass to the legend 
            d = dict(unique) # the tuples are converted to dictionaries with each label as the key, this is later used to filter legends for each room 
            
            unique1 = {key: value for key, value in d.items() if "1" in key} # the legends are filtered for each room and passed to the legend function
            leg1 = ax1.legend(unique1.values(), unique1.keys(),  loc='upper right', bbox_to_anchor=(0.79, 1), title = "$\mathbf{Child\/1}$")
            
            unique2 = {key: value for key, value in d.items() if "2" in key} # the legends are filtered for each room and passed to the legend function
            leg2 = ax1.legend(unique2.values(), unique2.keys(),  loc='upper right', bbox_to_anchor=(0.86, 1), title = "$\mathbf{Child\/2}$")
            
            unique3 = {key: value for key, value in d.items() if "3" in key} # the legends are filtered for each room and passed to the legend function
            leg3 = ax1.legend(unique3.values(), unique3.keys(), loc='upper right', bbox_to_anchor=(0.93, 1), title = "$\mathbf{Bedroom}$")
            
            unique4 = {key: value for key, value in d.items() if "4" in key} # the legends are filtered for each room and passed to the legend function
            leg4 = ax1.legend(unique4.values(), unique4.keys(), loc='upper right', bbox_to_anchor=(1, 1), title = "$\mathbf{Bathroom}$")
            
            ax1.add_artist(leg1) # everytime we introduce a legend the previous legend is replaced
            ax1.add_artist(leg2) # so leg1, leg2, leg3 are replaced by leg4
            ax1.add_artist(leg3) # we use add_artist to display the previously overridden legends
            
            # ax1.axvline(x = 0, c = 'k', ls = '--',lw = 0.5, label='Start of the Experiment')
            # ax1.axvline(x = 7440, c = 'k',ls = '--',lw = 0.5, label='End of the Experiment')
            # ax1.text(x = 10, y = -150, s = "Start of evaluation " , c = 'k')
            # ax1.text(x = 7470, y = -150, s = "End of evaluation", c = 'k')# this displays text on graph
            
            ax1.set_title('Abklingkurve comparison between pd SMA and log')
            #for i in plot_log_list:
            #    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax2)
            for i, j, k in list(zip(self.plot_log_list, self.end_runtime, self.plot_log_end_list)) :
                if "1L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax2, color = '#179C7D')
                    ax2.scatter(j, k, marker='o', color = '#179C7D')
                elif "2L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax2, color = '#F29400')
                    ax2.scatter(j, k, marker='o', color = '#F29400')
                elif "3L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax2, color = '#1F82C0')
                    ax2.scatter(j, k, marker='o', color = '#1F82C0')
                elif "4L" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime',ls = '--', ax = ax2, color = '#E2001A')
                    ax2.scatter(j, k, marker='o', color = '#E2001A')
                elif "1" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax2, color = '#179C7D')
                    ax2.scatter(j, k, marker='o', color = '#179C7D')
                elif "2" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax2, color = '#F29400')
                    ax2.scatter(j, k, marker='o', color = '#F29400')
                elif "3" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax2, color = '#1F82C0')
                    ax2.scatter(j, k, marker='o', color = '#1F82C0')
                elif "4" in i.columns[1]:
                    i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax2, color = '#E2001A')
                    ax2.scatter(j, k, marker='o', color = '#E2001A')    
            
            handles, labels = ax2.get_legend_handles_labels()
            unique = [(l, h) for i, (h, l) in enumerate(zip(handles, labels))]
            unique.sort(key = operator.itemgetter(0))
            d = dict(unique) 
            
            unique1 = {key: value for key, value in d.items() if "1" in key}
            leg1 = ax2.legend(unique1.values(), unique1.keys(),  loc='upper right', bbox_to_anchor=(0.79, 0.58), title = "$\mathbf{Child\/1}$")
            
            unique2 = {key: value for key, value in d.items() if "2" in key}
            leg2 = ax2.legend(unique2.values(), unique2.keys(),  loc='upper right', bbox_to_anchor=(0.86, 0.58), title = "$\mathbf{Child\/2}$")
            
            unique3 = {key: value for key, value in d.items() if "3" in key}
            leg3 = ax2.legend(unique3.values(), unique3.keys(), loc='upper right', bbox_to_anchor=(0.93, 0.58), title = "$\mathbf{Bedroom}$")
            
            unique4 = {key: value for key, value in d.items() if "4" in key}
            leg4 = ax2.legend(unique4.values(), unique4.keys(), loc='upper right', bbox_to_anchor=(1, 0.58), title = "$\mathbf{Bathroom}$")
            
            ax2.add_artist(leg1)
            ax2.add_artist(leg2)
            ax2.add_artist(leg3)
            
            # ax2.axvline(x = 0, c = 'k', ls = '--',lw = 0.5, label='Start of the Experiment')
            # ax2.axvline(x = 7440, c = 'k',ls = '--',lw = 0.5, label='End of the Experiment')
            # ax2.text(x = 10, y = -2, s = "Start of evaluation " , c = 'k')
            # ax2.text(x = 7470, y = -2, s = "End of evaluation", c = 'k')
            
            f.align_ylabels((ax1, ax2))         
            plt.xlabel("runtime t in s")
            ax2.set_ylabel("log")
            plt.savefig(path + "/" + 'decay_curve.png', bbox_inches='tight', dpi=400)
            
    #         runtime_end = max(self.end_runtime) + 1000
    #         plt.figure(figsize=(8,5))
    #         ax1 = plt.gca()
    #         plt.title("Abklingkurve all Delta_in-out_C")    
    #         for i, j, k in list(zip(self.plot_delta_list, self.end_runtime, self.plot_delta_end_list)):
    #             i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax1)
    #             ax1.scatter(j, k, marker='o')
    #         plt.ylim(top = self.c0()["C_CO2 in ppm"].iloc[0] + 500)
    #         plt.axvline(x = 0, c = 'g', ls = '--', label='Start of the Experiment')
    # #        plt.axvline(x = (self.j - self.i) * self.sec, c = 'r',ls = '--', label='End of the Experiment')
    #         plt.text(x = -510, y = 100, s = "Start of the Experiment: " + str(self.t0), c = 'g')
    # #        plt.text(x = (self.j - self.i) * self.sec, y = 100, s = "End of the Experiment: " + str(self.tend_es), c = 'r')
    #         plt.xlabel("runtime t in s")
    #         plt.ylabel("concentration offsett \Delta C_(CO_2) (t, x) in ppm")
    #         plt.legend(loc='upper right', ncol = 3)
    #         plt.show()
            
            
            
    #         plt.figure(figsize=(8,5))
    #         ax2 = plt.gca()
    #         plt.title("Abklingkurve all pd_SMA")    
    #         for i, j, k in list(zip(self.plot_pd_SMA_list, self.end_runtime, self.plot_pd_SMA_end_list)) :
    #             i[i['runtime']<runtime_end].plot(x = 'runtime', ax = ax2)
    #             ax2.scatter(j, k, marker='o')
    #         plt.ylim(top = self.c0()["C_CO2 in ppm"].iloc[0] + 500)
    #         plt.xlabel("runtime t in s")
    #         plt.ylabel("Simple Moving Average\Delta C_(CO_2) (t, x) in ppm")
    #         plt.axvline(x = 0, c = 'g', ls = '--', label='Start of the Experiment')
    # #        plt.axvline(x = (self.j - self.i) * self.sec, c = 'r', ls = '--', label='End of the Experiment')
    #         plt.text(x = -510, y = 100, s = "Start of the Experiment: " + str(self.t0), c = 'g')
    # #        plt.text(x = (self.j - self.i) * self.sec, y = 100, s = "End of the Experiment: " + str(self.tend_es), c = 'r')
    #         plt.xlabel("runtime t in s")
    #         plt.ylabel("concentration offsett \Delta C_(CO_2) (t, x) in ppm")
    #         plt.legend(loc='best', ncol = 3)
    #         plt.show()
            
            
            
    #%% Names
    names = pd.read_sql_query('SHOW TABLES FROM {}'.format(schema), engine) # gives a list of names of all tables in a schema
    exclude_always = ['bd_original', 'außen', 'weather', 'tr']
    
    exclude = exclude_always + times["exclude"][z].split(',')
    
    names = names[~names["Tables_in_{}".format(schema)].isin(exclude)].iloc[:,0]
    print("List of sensors analyzed\n{}".format(names))
    #%% To find the deviation in initial concentration
    
    initial_CO2 = []
    final_CO2 = []
    
    for i in names:
        tau = Tau_l(i)
        initial_CO2.append(tau.c0()["C_CO2 in ppm"][0])
        final_CO2.append(tau.cend()["C_CO2 in ppm"][0])
    s_c0 = statistics.stdev(initial_CO2)/mean(initial_CO2)    
    s_phi_e = statistics.stdev(final_CO2)/mean(final_CO2)    
    print(s_c0)
    #%%
    
    lis = []
    for i in names:
    
        tau = Tau_l(i)
        tau.uncertainity(s_c0)
    
        lis.append(tau)
    #c = Tau_l("1c")
    #c.uncertainity()
    #a = c.tau_l
    df_eps = tau.result()    
    tau.print_results()
      
    #%% Output Dataframe
result = pd.DataFrame(["Results"])
r = 1
result.loc[r,0] = "Database"
result.loc[r,1] = tau.database
r += 1
result.loc[r,0] = "Volume flow for this Experiment"
result.loc[r,1] = tau.Vdot

r += 1
result.loc[r,0] = "Level"
result.loc[r,1] = tau.level

info1 = info.T.reset_index(drop=False).T
result = pd.concat([result, info1], ignore_index=True)

r += 3
result.loc[r,0] = "Unblocked Volume"
result.loc[r,1] = tau.v
r += 1
result.loc[r,0] = "empty"
r += 1
result.loc[r,0] = "Average air age"
result.loc[r,1] = tau.tau_av
r += 1
result.loc[r,0] = "Global air exchange efficiency"
result.loc[r,1] = tau.ea
r += 1
result.loc[r,0] = "empty"
r += 1
result.loc[r,0] = 'Sensor wise results'
    
    df_eps1 = df_eps[0].reset_index().T.reset_index(drop=False).T
    result = pd.concat([result, df_eps1], ignore_index=True)
    
    
    
    
    #%%% save result
    
    tau.plot_test_2(path)
    
    
    path = "C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/results/results_final.xlsx"
    
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book
    
    result.to_excel(writer, index = False , sheet_name = times["short name"][experiment][:10])
    writer.save()
    writer.close()
    
    path = "C:/Users/Devineni/OneDrive - bwedu/MA_Raghavakrishna/0_Evaluation/results/df_eps.xlsx"
    
    book = load_workbook(path)
    writer = pd.ExcelWriter(path, engine = 'openpyxl')
    writer.book = book
    
    df_eps[0].reset_index().to_excel(writer, index = False , sheet_name = times["short name"][experiment][:10])
    writer.save()
    writer.close()




#%%
prYellow("#########################The End of Execution ###########################")    
prRed(database)
    
    